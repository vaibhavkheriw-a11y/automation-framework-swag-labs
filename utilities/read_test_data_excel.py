from openpyxl import load_workbook

def getRowCount(file, sheetName):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColumnCount(file, sheetName):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

def readData(file, sheetName, rownum, columnno):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file, sheetName, rownum, columnno, data):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)